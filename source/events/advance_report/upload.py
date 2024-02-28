from datetime import date, datetime
from io import BytesIO
from aiogram.types import CallbackQuery, BufferedInputFile
from aiogram_dialog import DialogManager, ShowMode
from aiogram_dialog.widgets.kbd import Button
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from tortoise.expressions import Q
from config import EXCEL_TEMPLATE_PATH, UPLOAD_AR_ONLY_BY_USER_LIST
from models import Project


async def on_report_person_selected(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    dialog_manager.dialog_data['report_person'] = button.text.__dict__['text']


async def on_date_start_selected(callback: CallbackQuery, widget,
                                 dialog_manager: DialogManager, selected_date: date):
    if callback.from_user.id in UPLOAD_AR_ONLY_BY_USER_LIST:
        dialog_manager.dialog_data['report_person'] = dialog_manager.start_data['report_person']
    dialog_manager.dialog_data["start_date"] = selected_date.__str__()
    await dialog_manager.next()


async def on_date_end_selected(callback: CallbackQuery, widget,
                               dialog_manager: DialogManager, selected_date: date):
    start_date = dialog_manager.dialog_data["start_date"]
    end_date = selected_date.__str__()
    report_person = dialog_manager.dialog_data['report_person']

    output = BytesIO()
    wb = load_workbook(filename=EXCEL_TEMPLATE_PATH)
    ws = wb.active

    # Собираем проекты по заданному интервалу --------------------------------------------------------------------------
    d_start_date = datetime.strptime(start_date, '%Y-%m-%d')
    d_end_date = datetime.strptime(end_date, '%Y-%m-%d')
    expression = (Q(advance_report__accountable_person=report_person) &
                  Q(advance_report__status=1) &
                  (Q(advance_report__datetime__range=[d_start_date, d_end_date]) |
                  Q(advance_report__datetime__startswith=start_date) |
                   Q(advance_report__datetime__startswith=end_date)))

    projects = await (Project.filter(expression).select_related('advance_report'))

    if not projects:
        await callback.message.answer("⛔️ За выбранный период, по подотчетному лицу нет ни одного "
                                      "согласованного авансового отчета.")
        await dialog_manager.done()
        return

    for i, project in enumerate(projects, start=2):
        ws[f'A{i}'] = datetime.strftime(project.advance_report.datetime, '%Y.%m.%d')
        ws[f'B{i}'] = project.project_name
        ws[f'C{i}'] = project.expense
        ws[f'D{i}'] = project.comment
        ws[f'E{i}'] = project.amount
        ws[f'F{i}'] = project.currency.value

        ws[f'D{i}'].alignment = Alignment(wrap_text=True)

    file_name = f"{report_person} АО (с {start_date} по {end_date}).xlsx"

    wb.save(output)

    await dialog_manager.next()
    await dialog_manager.show(show_mode=ShowMode.DELETE_AND_SEND)
    await callback.message.answer_document(document=BufferedInputFile(file=output.getvalue(), filename=file_name))
    await dialog_manager.done()
